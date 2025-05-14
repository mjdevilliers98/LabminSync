import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from ttkthemes import ThemedTk
import string
import serial
import serial.tools.list_ports
import threading
import time
import json
import os
import sys
import requests
import datetime
import csv
import openpyxl
import PyPDF2
import pyautogui
import platform
import getpass
import traceback
import base64
import uuid
import subprocess
import xlrd
import xml.etree.ElementTree as ET
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import queue
import logging
import shutil
import hashlib
import zipfile

# HELPER: CREATE CONFIG PATH
def config_path(filename):
    config_dir = os.path.expanduser("~/.serial_app")
    os.makedirs(config_dir, exist_ok=True)
    return os.path.join(config_dir, filename)

# MAIN CLASS
class SerialApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Labmin Sync – Professional Edition v2.0")

        # Setup logging
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('labmin_test_app.log'),
                logging.StreamHandler()
            ]
        )
        logger = logging.getLogger(__name__)

        # Paths for config files
        self.devices_config_path = config_path("devices_config.json")
        self.webhook_config_path = config_path("webhook_config.json")
        self.settings_config_path = config_path("settings.json")
        self.license_config_path = config_path("license.json")
        self.main_log_file = config_path("labmin_logs.txt")

        # System info for license
        self.system_mac = self.get_mac_address()
        self.system_serial = self.get_serial_number()

        # Load configurations
        self.load_configurations()

        # Check license
        if not self.verify_license():
            messagebox.showerror("License Error", "Invalid or missing license key. The application will now exit.")
            root.destroy()
            return

        # Internals
        self.device_tabs = []
        self.serial_ports_in_use = set()
        self.serial_ports = []
        self.reading_threads = []
        self.reading_flags = []
        self.monitoring_threads = []
        self.monitoring_flags = []
        self.observers = []
        self.heartbeat_flag = not self.settings_config.get('offline_mode', False)
        self.data_queues = []
        self.webhook_threads = []
        self.token = None
        self.token_expiry = 0
        self.token_lock = threading.Lock()

        # Create style + menubar
        self.create_style()
        self.create_menubar()

        # Main GUI
        self.create_gui()

        # Start minimized if configured
        if self.settings_config.get('launch_minimized', False):
            self.root.iconify()

        # Auto-size to content, then fix a minimal size
        self.root.update_idletasks()
        w = self.root.winfo_reqwidth()
        h = self.root.winfo_reqheight()
        self.root.minsize(w, h)

        # Request token if not in offline mode
        if not self.settings_config.get('offline_mode', False):
            self.request_token()

        # Start devices and heartbeat
        self.root.after(1000, self.start_all_devices)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.after(5000, self.refresh_ports_periodically)

    ### LICENSE
    def verify_license(self):
        license_data = self.load_license()
        license_key = license_data.get("license_key", "")
        offline_mode = self.settings_config.get('offline_mode', False)

        if not license_key:
            license_key = simpledialog.askstring(
                "License Key",
                "Enter license key (required even in offline mode):",
                show='*'
            )
            if not license_key:
                return False

        if not self.system_mac and not self.system_serial:
            messagebox.showerror("Error", "Unable to retrieve system info for license verification.")
            return False

        salt = "Labmin2025"
        local_hash = hashlib.sha256(
            f"{license_key}{self.system_mac}{self.system_serial}{salt}".encode()
        ).hexdigest()

        if offline_mode:
            if license_key:
                license_data["license_key"] = license_key
                license_data["local_hash"] = local_hash
                license_data["last_verified"] = datetime.datetime.now().isoformat()
                self.save_license(license_data)
                return True
            return False

        server_valid = self.check_license_with_server(license_key, self.system_mac, self.system_serial)
        if server_valid:
            license_data = {
                "license_key": license_key,
                "mac_address": self.system_mac,
                "serial_number": self.system_serial,
                "local_hash": local_hash,
                "last_verified": datetime.datetime.now().isoformat()
            }
            self.save_license(license_data)
            return True
        else:
            stored_hash = license_data.get("local_hash", "")
            last_verified = license_data.get("last_verified", "")
            if stored_hash == local_hash and last_verified and (
                datetime.datetime.now() - datetime.datetime.fromisoformat(last_verified)
            ).days < 30:
                logging.info("License validated offline using local hash.")
                return True
            messagebox.showerror("License Error", "License verification failed.")
            return False

    def load_license(self):
        if os.path.exists(self.license_config_path):
            with open(self.license_config_path, 'r') as f:
                return json.load(f)
        return {}

    def save_license(self, license_data):
        with open(self.license_config_path, 'w') as f:
            json.dump(license_data, f, indent=4)

    def get_mac_address(self):
        try:
            mac_num = hex(uuid.getnode()).replace('0x', '').upper()
            return ':'.join(mac_num[i:i+2] for i in range(0, 11, 2))
        except Exception as e:
            logging.error(f"Error getting MAC: {e}")
            return None

    def get_serial_number(self):
        os_name = platform.system()
        try:
            if os_name == 'Windows':
                cmd = 'wmic bios get serialnumber'
                output = subprocess.check_output(cmd, shell=True).decode()
                return output.strip().split('\n')[-1].strip()
            elif os_name == 'Linux':
                cmd = "sudo dmidecode -s system-serial-number"
                output = subprocess.check_output(cmd, shell=True).decode()
                return output.strip()
            elif os_name == 'Darwin':
                cmd = "ioreg -l | grep IOPlatformSerialNumber"
                output = subprocess.check_output(cmd, shell=True).decode()
                return output.split('"')[-2]
            return None
        except Exception as e:
            logging.error(f"Error getting serial number: {e}")
            return None

    def check_license_with_server(self, license_key, mac_address, serial_number):
        if self.settings_config.get('offline_mode', False):
            return False
        payload = {"license_key": license_key}
        if mac_address:
            payload["mac_address"] = mac_address
        if serial_number:
            payload["serial_number"] = serial_number
        try:
            r = requests.post("https://iot.labmin.mobi/verify_license", json=payload, timeout=5)
            return r.status_code == 200
        except Exception as e:
            logging.error(f"License server check failed: {e}")
            return False

    ### STYLE, MENUBAR
    def create_style(self):
        style = ttk.Style(self.root)
        style.theme_use("adapta")
        style.configure("TLabel", font=("Segoe UI", 8))
        style.configure("TButton", font=("Segoe UI", 8))
        style.configure("TEntry", font=("Segoe UI", 8))
        style.configure("TNotebook.Tab", font=("Segoe UI", 8))
        style.configure("TFrame", background="#f8f8f8")
        style.configure("TLabelframe.Label", font=("Segoe UI", 8, "bold"))

    def create_menubar(self):
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=False)
        file_menu.add_command(label="Exit", command=self.menu_exit)
        menubar.add_cascade(label="File", menu=file_menu)

        settings_menu = tk.Menu(menubar, tearoff=False)
        settings_menu.add_command(label="Configure", command=self.open_settings_window)
        menubar.add_cascade(label="Settings", menu=settings_menu)

        connection_menu = tk.Menu(menubar, tearoff=False)
        connection_menu.add_command(
            label="Test Connection",
            command=self.open_test_connection_window,
            state='disabled' if self.settings_config.get('offline_mode', False) else 'normal'
        )
        menubar.add_cascade(label="Connection", menu=connection_menu)

        tools_menu = tk.Menu(menubar, tearoff=False)
        tools_menu.add_command(label="Export Logs", command=self.export_logs)
        tools_menu.add_command(label="Backup Config", command=self.backup_config)
        tools_menu.add_command(label="Restore Config", command=self.restore_config)
        menubar.add_cascade(label="Tools", menu=tools_menu)

        help_menu = tk.Menu(menubar, tearoff=False)
        help_menu.add_command(
            label="About",
            command=lambda: messagebox.showinfo(
                "About", "Labmin Sync Professional v2.0\n© 2025 Labmin"
            )
        )
        menubar.add_cascade(label="Help", menu=help_menu)

        self.root.config(menu=menubar)

    def menu_exit(self):
        self.on_closing()
        sys.exit(0)

    ### LOAD CONFIG
    def load_configurations(self):
        if os.path.exists(self.devices_config_path):
            with open(self.devices_config_path, 'r') as f:
                self.devices_config = json.load(f)
        else:
            self.devices_config = {"devices": [{"name": "Device1", "config_file": "device1_config.json"}]}
            self.save_config(self.devices_config_path, self.devices_config)

        if os.path.exists(self.webhook_config_path):
            with open(self.webhook_config_path, 'r') as f:
                self.webhook_config = json.load(f)
        else:
            self.webhook_config = {
                "webhook_url": "https://iot.labmin.mobi",
                "heartbeat_webhook_url": "https://iot.labmin.mobi",
                "token_url": "https://iot.labmin.mobi/get_token",
                "webhook_headers": {"Content-Type": "application/json"}
            }
            self.save_config(self.webhook_config_path, self.webhook_config)

        if os.path.exists(self.settings_config_path):
            with open(self.settings_config_path, 'r') as f:
                self.settings_config = json.load(f)
        else:
            self.settings_config = {
                "launch_on_startup": False,
                "launch_minimized": False,
                "computer_name": "MyComputer",
                "offline_mode": False,
                "printer_settings": {
                    "enabled": False,
                    "baudrate": "9600",
                    "init_char": "@",
                    "expected_response": "$(O)$",
                    "command": "R1",
                    "end_char": "#",
                    "reset_connection": True,
                    "webhook_filter_min": "0",
                    "webhook_filter_max": "100",
                    "webhook_filter_enabled": False
                }
            }
            self.save_config(self.settings_config_path, self.settings_config)

    def save_config(self, path, config):
        with open(path, 'w') as f:
            json.dump(config, f, indent=4)

    ### MAIN GUI
    def create_gui(self):
        banner_frame = ttk.Frame(self.root, padding=5)
        banner_frame.pack(fill='x', side='top')
        banner_label = ttk.Label(
            banner_frame,
            text="Labmin Sync – Professional Edition v2.0",
            font=("Segoe UI", 10, "bold")
        )
        banner_label.pack(side="left", padx=5)

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=5, pady=5, side='top')
        self.create_device_tabs()

        save_all_frame = ttk.Frame(self.root, padding=5)
        save_all_frame.pack(fill='x', side='top')
        save_btn = ttk.Button(
            save_all_frame,
            text="Save All Device Settings",
            command=self.save_all_device_configs
        )
        save_btn.pack(side='right')

        self.status_var = tk.StringVar(value="Active Devices: 0")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief='sunken', anchor='w')
        status_bar.pack(fill='x', side='bottom', padx=5, pady=2)

    def create_device_tabs(self):
        for i, dev in enumerate(self.devices_config['devices']):
            dev_tab = ttk.Frame(self.notebook, padding=5)
            self.notebook.add(dev_tab, text=dev['name'])
            self.device_tabs.append(dev_tab)
            self.data_queues.append(queue.Queue())
            self.root.after(100, lambda t=dev_tab, idx=i, d=dev: self.create_device_tab(t, idx, d))

    def create_device_tab(self, tab, device_index, device_info):
        dev_file = config_path(device_info['config_file'])
        dev_config = json.load(open(dev_file, 'r')) if os.path.exists(dev_file) else {}
        token_locked = dev_config.get('token_locked', False)

        self.serial_ports.append(None)
        self.reading_threads.append(None)
        self.reading_flags.append(False)
        self.monitoring_threads.append(None)
        self.monitoring_flags.append(False)

        container = ttk.Frame(tab)
        container.pack(expand=True, fill='both', side='top')

        left_frame = ttk.Frame(container)
        left_frame.pack(side='left', fill='y', expand=False, padx=3, pady=3)
        right_frame = ttk.Frame(container)
        right_frame.pack(side='right', fill='both', expand=True, padx=3, pady=3)

        heartbeat_indicator = tk.Label(left_frame, text="   ", bg='red', width=3, height=1)
        heartbeat_indicator.pack(anchor='ne', pady=(0,5))

        dev_frame = ttk.Labelframe(left_frame, text="Device Setup", padding=5)
        dev_frame.pack(fill='x', padx=2, pady=2)

        ttk.Label(dev_frame, text="Device Name:").grid(row=0, column=0, sticky='w', padx=2, pady=2)
        dev_name_var = tk.StringVar(value=device_info.get('name', f"Device{device_index+1}"))
        dev_name_entry = ttk.Entry(dev_frame, textvariable=dev_name_var, width=20)
        dev_name_entry.grid(row=0, column=1, sticky='w', padx=2, pady=2)

        ttk.Label(dev_frame, text="Data Type:").grid(row=1, column=0, sticky='w', padx=2, pady=2)
        data_type_var = tk.StringVar(value=dev_config.get('data_type', 'txt'))
        dt_frame = ttk.Frame(dev_frame)
        dt_frame.grid(row=1, column=1, sticky='w', padx=2, pady=2)
        data_types = ['txt', 'csv', 'xls', 'pdf', 'printer']
        dt_rbs = {}
        for dt_ in data_types:
            label_ = dt_.upper() if dt_ != 'printer' else 'PRINTER'
            rb = ttk.Radiobutton(dt_frame, text=label_, variable=data_type_var, value=dt_)
            rb.pack(side='left', padx=2)
            dt_rbs[dt_] = rb
        if not self.settings_config.get('printer_settings', {}).get('enabled', False):
            dt_rbs['printer'].pack_forget()  # Hide the printer radio button

        ttk.Label(dev_frame, text="Handling:").grid(row=2, column=0, sticky='w', padx=2, pady=2)
        handling_var = tk.StringVar(value=dev_config.get('handling_method', 'webhook'))
        handling_frame = ttk.Frame(dev_frame)
        handling_frame.grid(row=2, column=1, sticky='w', padx=2, pady=2)
        handling_methods = [('Webhook', 'webhook'), ('Paste to Cursor', 'cursor')]
        handling_rbs = {}
        for tlabel, val in handling_methods:
            rb = ttk.Radiobutton(handling_frame, text=tlabel, variable=handling_var, value=val)
            rb.pack(side='left', padx=2)
            handling_rbs[val] = rb

        ttk.Label(dev_frame, text="Folder:").grid(row=3, column=0, sticky='w', padx=2, pady=2)
        folder_path_var = tk.StringVar(value=dev_config.get('folder_path', ''))
        folder_fr = ttk.Frame(dev_frame)
        folder_fr.grid(row=3, column=1, sticky='w', padx=2, pady=2)
        folder_entry = ttk.Entry(folder_fr, textvariable=folder_path_var, state='disabled', width=16)
        folder_entry.pack(side='left', fill='x')
        browse_btn = ttk.Button(
            folder_fr,
            text="...",
            width=3,
            state='disabled',
            command=lambda: self.browse_folder(folder_path_var)
        )
        browse_btn.pack(side='left', padx=2)

        ttk.Label(dev_frame, text="Token:").grid(row=4, column=0, sticky='w', padx=2, pady=2)
        token_var = tk.StringVar(value=dev_config.get('token', ''))
        token_entry_state = 'disabled' if token_locked else 'normal'
        token_entry = ttk.Entry(dev_frame, textvariable=token_var, width=16, state=token_entry_state)
        token_entry.grid(row=4, column=1, sticky='w', padx=2, pady=2)

        lock_button = ttk.Button(
            dev_frame,
            text="🔒" if token_locked else "🔓",
            width=3,
            command=lambda: self.toggle_token_lock(device_index)
        )
        lock_button.grid(row=4, column=2, sticky='w', padx=2)

        serial_fr = ttk.Labelframe(left_frame, text="Serial Config", padding=5)
        serial_fr.pack(fill='x', padx=2, pady=2)

        ttk.Label(serial_fr, text="Port:").grid(row=0, column=0, sticky='w', padx=2, pady=2)
        port_var = tk.StringVar(value=dev_config.get('port', ''))
        port_combo = ttk.Combobox(serial_fr, textvariable=port_var, values=self.get_ports(), width=10)
        port_combo.grid(row=0, column=1, sticky='w', padx=2, pady=2)

        ttk.Label(serial_fr, text="Baud:").grid(row=1, column=0, sticky='w', padx=2, pady=2)
        baud_var = tk.StringVar(value=dev_config.get('baudrate', '9600'))
        baud_combo = ttk.Combobox(
            serial_fr,
            textvariable=baud_var,
            values=["9600", "19200", "38400", "57600", "115200"],
            width=10
        )
        baud_combo.grid(row=1, column=1, sticky='w', padx=2, pady=2)

        ttk.Label(serial_fr, text="Parity:").grid(row=2, column=0, sticky='w', padx=2, pady=2)
        parity_var = tk.StringVar(value=dev_config.get('parity', 'None'))
        parity_combo = ttk.Combobox(
            serial_fr,
            textvariable=parity_var,
            values=["None", "Even", "Odd", "Mark", "Space"],
            width=10
        )
        parity_combo.grid(row=2, column=1, sticky='w', padx=2, pady=2)

        ttk.Label(serial_fr, text="Data Bits:").grid(row=3, column=0, sticky='w', padx=2, pady=2)
        data_bits_var = tk.StringVar(value=dev_config.get('data_bits', '8'))
        data_bits_combo = ttk.Combobox(
            serial_fr,
            textvariable=data_bits_var,
            values=["5", "6", "7", "8"],
            width=10
        )
        data_bits_combo.grid(row=3, column=1, sticky='w', padx=2, pady=2)

        ttk.Label(serial_fr, text="Stop Bits:").grid(row=4, column=0, sticky='w', padx=2, pady=2)
        stop_bits_var = tk.StringVar(value=dev_config.get('stop_bits', '1'))
        stop_bits_combo = ttk.Combobox(
            serial_fr,
            textvariable=stop_bits_var,
            values=["1", "1.5", "2"],
            width=10
        )
        stop_bits_combo.grid(row=4, column=1, sticky='w', padx=2, pady=2)

        ttk.Label(serial_fr, text="Flow Ctrl:").grid(row=5, column=0, sticky='w', padx=2, pady=2)
        flow_control_var = tk.StringVar(value=dev_config.get('flow_control', 'None'))
        flow_control_combo = ttk.Combobox(
            serial_fr,
            textvariable=flow_control_var,
            values=["None", "RTS/CTS", "XON/XOFF"],
            width=10
        )
        flow_control_combo.grid(row=5, column=1, sticky='w', padx=2, pady=2)

        ttk.Label(serial_fr, text="Interval (s):").grid(row=6, column=0, sticky='w', padx=2, pady=2)
        interval_var = tk.StringVar(value=dev_config.get('interval', '1'))
        interval_entry = ttk.Entry(serial_fr, textvariable=interval_var, width=6)
        interval_entry.grid(row=6, column=1, sticky='w', padx=2, pady=2)

        actions_frame = ttk.Frame(left_frame)
        actions_frame.pack(fill='x', padx=2, pady=5)

        start_btn = ttk.Button(
            actions_frame,
            text="Start",
            command=lambda: self.start_device(device_index)
        )
        start_btn.pack(side='left', padx=2)

        stop_btn = ttk.Button(
            actions_frame,
            text="Stop",
            state='disabled',
            command=lambda: self.stop_device(device_index)
        )
        stop_btn.pack(side='left', padx=2)

        save_dev_btn = ttk.Button(
            actions_frame,
            text="Save Device",
            command=lambda: self.save_single_device_config(device_index, quiet=False)
        )
        save_dev_btn.pack(side='left', padx=10)

        status_lbl = ttk.Label(actions_frame, text="Status: Disconnected", foreground='red')
        status_lbl.pack(side='left', padx=10)

        log_fr = ttk.Labelframe(right_frame, text="Log Output", padding=5)
        log_fr.pack(fill='both', expand=True)

        text_box = tk.Text(log_fr, wrap='word', height=10, width=50)
        text_box.pack(fill='both', expand=True, side='top')
        text_box.config(state='disabled', font=("Segoe UI", 8))

        clear_log_btn = ttk.Button(
            right_frame,
            text="Clear Log",
            command=lambda: self.clear_log(text_box)
        )
        clear_log_btn.pack(pady=2)

        stats_frame = ttk.Labelframe(right_frame, text="Statistics", padding=5)
        stats_frame.pack(fill='x', pady=5)
        bytes_received_var = tk.StringVar(value="Bytes: 0")
        ttk.Label(stats_frame, textvariable=bytes_received_var).pack(anchor='w')

        tab.widgets = {
            'device_info': device_info,
            'heartbeat_indicator': heartbeat_indicator,
            'device_name': dev_name_var,
            'device_name_entry': dev_name_entry,
            'data_type_var': data_type_var,
            'data_type_radiobuttons': dt_rbs,
            'handling_var': handling_var,
            'handling_radiobuttons': handling_rbs,
            'folder_path_var': folder_path_var,
            'folder_entry': folder_entry,
            'browse_button': browse_btn,
            'token_var': token_var,
            'token_entry': token_entry,
            'token_locked': token_locked,
            'token_lock_button': lock_button,
            'port_var': port_var,
            'port_combo': port_combo,
            'baud_var': baud_var,
            'baud_combo': baud_combo,
            'parity_var': parity_var,
            'parity_combo': parity_combo,
            'data_bits_var': data_bits_var,
            'data_bits_combo': data_bits_combo,
            'stop_bits_var': stop_bits_var,
            'stop_bits_combo': stop_bits_combo,
            'flow_control_var': flow_control_var,
            'flow_control_combo': flow_control_combo,
            'interval_var': interval_var,
            'interval_entry': interval_entry,
            'start_button': start_btn,
            'stop_button': stop_btn,
            'status_label': status_lbl,
            'text_box': text_box,
            'bytes_received_var': bytes_received_var
        }

        data_type_var.trace_add('write', lambda *a: self.update_handling_options(tab))
        handling_var.trace_add('write', lambda *a: self.update_data_type_options(tab))
        self.update_handling_options(tab)
        self.update_data_type_options(tab)
        if self.settings_config.get('offline_mode', False):
            self.enforce_offline_mode(tab)

    ### SAVE SINGLE DEVICE
    def save_single_device_config(self, dev_index, quiet=True):
        tab = self.device_tabs[dev_index]
        w = tab.widgets
        dev_info = w['device_info']

        dev_name = w['device_name'].get()
        dev_conf_file = config_path(dev_info['config_file'])
        dev_conf = {
            'name': dev_name,
            'token': w['token_var'].get(),
            'token_locked': w['token_locked'],
            'data_type': w['data_type_var'].get(),
            'handling_method': w['handling_var'].get(),
            'folder_path': w['folder_path_var'].get(),
            'port': w['port_var'].get(),
            'baudrate': w['baud_var'].get(),
            'parity': w['parity_var'].get(),
            'data_bits': w['data_bits_var'].get(),
            'stop_bits': w['stop_bits_var'].get(),
            'flow_control': w['flow_control_var'].get(),
            'interval': w['interval_var'].get()
        }
        self.devices_config['devices'][dev_index]['name'] = dev_name
        self.save_config(dev_conf_file, dev_conf)
        self.save_config(self.devices_config_path, self.devices_config)
        if not quiet:
            messagebox.showinfo("Saved", f"Device '{dev_name}' settings saved successfully.")

    ### CONNECTION TEST WINDOW
    def open_test_connection_window(self):
        if self.settings_config.get('offline_mode', False):
            messagebox.showinfo("Offline Mode", "Test Connection is disabled in offline mode.")
            return
        conn_win = tk.Toplevel(self.root)
        conn_win.title("Test Connection")
        conn_win.transient(self.root)
        conn_win.grab_set()

        frame = ttk.Frame(conn_win, padding=10)
        frame.pack(expand=True, fill='both')

        ttk.Label(frame, text="Webhook URL:").grid(row=0, column=0, sticky='w', padx=3, pady=3)
        self.test_webhook_label_var = tk.StringVar(value=self.webhook_config.get('webhook_url', ''))
        ttk.Label(frame, textvariable=self.test_webhook_label_var).grid(row=0, column=1, sticky='w')

        webhook_btn = ttk.Button(frame, text="Refresh", command=self.test_main_webhook_connection)
        webhook_btn.grid(row=1, column=0, columnspan=2, pady=2, sticky='w')

        self.test_webhook_status_var = tk.StringVar(value="Checking...")
        self.test_webhook_status_label = ttk.Label(
            frame,
            textvariable=self.test_webhook_status_var,
            foreground='black'
        )
        self.test_webhook_status_label.grid(row=2, column=0, columnspan=2, pady=5, sticky='w')

        ttk.Label(frame, text="Heartbeat URL:").grid(row=3, column=0, sticky='w', padx=3, pady=3)
        self.test_heartbeat_label_var = tk.StringVar(
            value=self.webhook_config.get('heartbeat_webhook_url', '')
        )
        ttk.Label(frame, textvariable=self.test_heartbeat_label_var).grid(row=3, column=1, sticky='w')

        heartbeat_btn = ttk.Button(frame, text="Refresh", command=self.test_heartbeat_connection)
        heartbeat_btn.grid(row=4, column=0, columnspan=2, pady=2, sticky='w')

        self.test_heartbeat_status_var = tk.StringVar(value="Checking...")
        self.test_heartbeat_status_label = ttk.Label(
            frame,
            textvariable=self.test_heartbeat_status_var,
            foreground='black'
        )
        self.test_heartbeat_status_label.grid(row=5, column=0, columnspan=2, pady=5, sticky='w')

        close_frame = ttk.Frame(frame)
        close_frame.grid(row=6, column=0, columnspan=2, pady=10, sticky='e')
        ttk.Button(close_frame, text="Close", command=conn_win.destroy).pack(side='right', padx=5)

        self.root.after(100, self.test_main_webhook_connection)
        self.root.after(200, self.test_heartbeat_connection)

    def test_main_webhook_connection(self):
        url = self.webhook_config.get('webhook_url', '')
        headers = self.webhook_config.get('webhook_headers', {})
        try:
            payload = {"test": "connection"}
            r = requests.post(url, json=payload, headers=headers, timeout=5)
            if r.status_code == 200:
                self.test_webhook_status_var.set("Connected")
                self.test_webhook_status_label.config(foreground='green')
            else:
                self.test_webhook_status_var.set(f"Disconnected ({r.status_code})")
                self.test_webhook_status_label.config(foreground='red')
        except Exception as e:
            self.test_webhook_status_var.set(f"Disconnected ({str(e)})")
            self.test_webhook_status_label.config(foreground='red')

    def test_heartbeat_connection(self):
        url = self.webhook_config.get('heartbeat_webhook_url', '')
        headers = self.webhook_config.get('webhook_headers', {})
        try:
            payload = {"test": "heartbeat"}
            r = requests.post(url, json=payload, headers=headers, timeout=5)
            if r.status_code == 200:
                self.test_heartbeat_status_var.set("Connected")
                self.test_heartbeat_status_label.config(foreground='green')
            else:
                self.test_heartbeat_status_var.set(f"Disconnected ({r.status_code})")
                self.test_heartbeat_status_label.config(foreground='red')
        except Exception as e:
            self.test_heartbeat_status_var.set(f"Disconnected ({str(e)})")
            self.test_heartbeat_status_label.config(foreground='red')

    ### SETTINGS WINDOW
    def open_settings_window(self):
        settings_win = tk.Toplevel(self.root)
        settings_win.title("Settings")
        settings_win.transient(self.root)
        settings_win.grab_set()

        frame = ttk.Frame(settings_win, padding=10)
        frame.pack(expand=True, fill='both')

        notebook = ttk.Notebook(frame)
        notebook.pack(fill='both', expand=True, padx=5, pady=5)

        # Webhook Settings Tab
        webhook_tab = ttk.Frame(notebook)
        notebook.add(webhook_tab, text="Webhook")
        webhook_frame = ttk.Labelframe(webhook_tab, text="Webhook Settings", padding=5)
        webhook_frame.pack(fill='x', padx=5, pady=5)

        ttk.Label(webhook_frame, text="Data Webhook URL:").grid(row=0, column=0, sticky='w', padx=3, pady=3)
        data_webhook_var = tk.StringVar(value=self.webhook_config.get('webhook_url', ''))
        data_webhook_entry = ttk.Entry(webhook_frame, textvariable=data_webhook_var, width=40)
        data_webhook_entry.grid(row=0, column=1, sticky='ew', padx=3, pady=3)

        ttk.Label(webhook_frame, text="Heartbeat Webhook URL:").grid(row=1, column=0, sticky='w', padx=3, pady=3)
        heartbeat_url_var = tk.StringVar(value=self.webhook_config.get('heartbeat_webhook_url', ''))
        heartbeat_entry = ttk.Entry(webhook_frame, textvariable=heartbeat_url_var, width=40)
        heartbeat_entry.grid(row=1, column=1, sticky='ew', padx=3, pady=3)

        ttk.Label(webhook_frame, text="Token URL:").grid(row=2, column=0, sticky='w', padx=3, pady=3)
        token_url_var = tk.StringVar(value=self.webhook_config.get('token_url', ''))
        token_url_entry = ttk.Entry(webhook_frame, textvariable=token_url_var, width=40)
        token_url_entry.grid(row=2, column=1, sticky='ew', padx=3, pady=3)

        save_webhook_btn = ttk.Button(
            webhook_frame,
            text="Save",
            command=lambda: self.save_webhook_config(data_webhook_var, heartbeat_url_var, token_url_var)
        )
        save_webhook_btn.grid(row=3, column=1, sticky='e', padx=3, pady=3)

        # Computer Identification Tab
        comp_tab = ttk.Frame(notebook)
        notebook.add(comp_tab, text="Computer")
        comp_frame = ttk.Labelframe(comp_tab, text="Computer Identification", padding=5)
        comp_frame.pack(fill='x', padx=5, pady=5)

        ttk.Label(comp_frame, text="Computer Name:").grid(row=0, column=0, sticky='w', padx=3, pady=3)
        comp_name_var = tk.StringVar(value=self.settings_config.get('computer_name', 'MyComputer'))
        comp_name_entry = ttk.Entry(comp_frame, textvariable=comp_name_var, width=30)
        comp_name_entry.grid(row=0, column=1, sticky='ew', padx=3, pady=3)

        save_comp_btn = ttk.Button(
            comp_frame,
            text="Save Computer Name",
            command=lambda: self.save_computer_name(comp_name_var)
        )
        save_comp_btn.grid(row=0, column=2, sticky='e', padx=3, pady=3)

        # Manage Devices Tab
        dev_mgmt_tab = ttk.Frame(notebook)
        notebook.add(dev_mgmt_tab, text="Devices")
        dev_mgmt_frame = ttk.Labelframe(dev_mgmt_tab, text="Manage Devices", padding=5)
        dev_mgmt_frame.pack(fill='x', padx=5, pady=5)

        add_btn = ttk.Button(dev_mgmt_frame, text="Add Device", command=self.add_device)
        add_btn.grid(row=0, column=0, sticky='w', padx=5, pady=3)

        ttk.Label(dev_mgmt_frame, text="Remove Device:").grid(row=0, column=1, sticky='w', padx=5)
        self.remove_combo_var = tk.StringVar()
        self.remove_combo = ttk.Combobox(dev_mgmt_frame, textvariable=self.remove_combo_var, width=15)
        self.remove_combo.grid(row=0, column=2, sticky='w')
        self.populate_remove_device_combo()

        remove_btn = ttk.Button(dev_mgmt_frame, text="Remove", command=self.remove_selected_device)
        remove_btn.grid(row=0, column=3, sticky='w', padx=5)

        # Application Settings Tab
        app_sett_tab = ttk.Frame(notebook)
        notebook.add(app_sett_tab, text="Application")
        app_sett_frame = ttk.Labelframe(app_sett_tab, text="Application Settings", padding=5)
        app_sett_frame.pack(fill='x', padx=5, pady=5)

        launch_on_start_var = tk.BooleanVar(value=self.settings_config.get('launch_on_startup', False))
        launch_start_check = ttk.Checkbutton(
            app_sett_frame,
            text="Launch app on startup",
            variable=launch_on_start_var,
            command=lambda: self.toggle_launch_on_startup(launch_on_start_var)
        )
        launch_start_check.pack(anchor='w', pady=2)

        launch_min_var = tk.BooleanVar(value=self.settings_config.get('launch_minimized', False))
        launch_min_check = ttk.Checkbutton(
            app_sett_frame,
            text="Launch app minimized",
            variable=launch_min_var,
            command=lambda: self.toggle_launch_minimized(launch_min_var)
        )
        launch_min_check.pack(anchor='w', pady=2)

        offline_mode_var = tk.BooleanVar(value=self.settings_config.get('offline_mode', False))
        offline_mode_check = ttk.Checkbutton(
            app_sett_frame,
            text="Use App Offline Mode",
            variable=offline_mode_var,
            command=lambda: self.toggle_offline_mode(offline_mode_var, settings_win)
        )
        offline_mode_check.pack(anchor='w', pady=2)

        # Printer Settings Tab
        printer_tab = ttk.Frame(notebook)
        notebook.add(printer_tab, text="Printer")
        printer_frame = ttk.Labelframe(printer_tab, text="Printer Settings", padding=5)
        printer_frame.pack(fill='x', padx=5, pady=5)

        printer_enabled_var = tk.BooleanVar(value=self.settings_config.get('printer_settings', {}).get('enabled', False))
        printer_enabled_check = ttk.Checkbutton(
            printer_frame,
            text="Enable Printer",
            variable=printer_enabled_var,
            command=lambda: self.toggle_printer_settings(printer_enabled_var, printer_frame)
        )
        printer_enabled_check.grid(row=0, column=0, columnspan=2, sticky='w', padx=3, pady=3)

        ttk.Label(printer_frame, text="Baud Rate:").grid(row=1, column=0, sticky='w', padx=3, pady=3)
        printer_baud_var = tk.StringVar(value=self.settings_config.get('printer_settings', {}).get('baudrate', '9600'))
        printer_baud_combo = ttk.Combobox(
            printer_frame,
            textvariable=printer_baud_var,
            values=["9600", "19200", "38400", "57600", "115200"],
            width=10
        )
        printer_baud_combo.grid(row=1, column=1, sticky='w', padx=3, pady=3)

        ttk.Label(printer_frame, text="Initiation Character:").grid(row=2, column=0, sticky='w', padx=3, pady=3)
        init_char_var = tk.StringVar(value=self.settings_config.get('printer_settings', {}).get('init_char', '@'))
        init_char_entry = ttk.Entry(printer_frame, textvariable=init_char_var, width=5)
        init_char_entry.grid(row=2, column=1, sticky='w', padx=3, pady=3)

        ttk.Label(printer_frame, text="Expected Response:").grid(row=3, column=0, sticky='w', padx=3, pady=3)
        expected_response_var = tk.StringVar(value=self.settings_config.get('printer_settings', {}).get('expected_response', '$(O)$'))
        expected_response_entry = ttk.Entry(printer_frame, textvariable=expected_response_var, width=10)
        expected_response_entry.grid(row=3, column=1, sticky='w', padx=3, pady=3)

        ttk.Label(printer_frame, text="Command to Send:").grid(row=4, column=0, sticky='w', padx=3, pady=3)
        command_var = tk.StringVar(value=self.settings_config.get('printer_settings', {}).get('command', 'R1'))
        command_entry = ttk.Entry(printer_frame, textvariable=command_var, width=5)
        command_entry.grid(row=4, column=1, sticky='w', padx=3, pady=3)

        ttk.Label(printer_frame, text="End Character:").grid(row=5, column=0, sticky='w', padx=3, pady=3)
        end_char_var = tk.StringVar(value=self.settings_config.get('printer_settings', {}).get('end_char', '#'))
        end_char_entry = ttk.Entry(printer_frame, textvariable=end_char_var, width=5)
        end_char_entry.grid(row=5, column=1, sticky='w', padx=3, pady=3)

        reset_connection_var = tk.BooleanVar(value=self.settings_config.get('printer_settings', {}).get('reset_connection', True))
        reset_connection_check = ttk.Checkbutton(
            printer_frame,
            text="Reset connection after data received",
            variable=reset_connection_var
        )
        reset_connection_check.grid(row=6, column=0, columnspan=2, sticky='w', padx=3, pady=3)

        ttk.Label(printer_frame, text="Webhook Filter:").grid(row=7, column=0, sticky='w', padx=3, pady=3)
        webhook_filter_enabled_var = tk.BooleanVar(value=self.settings_config.get('printer_settings', {}).get('webhook_filter_enabled', False))
        webhook_filter_check = ttk.Checkbutton(
            printer_frame,
            text="Do not send to webhook if value between:",
            variable=webhook_filter_enabled_var
        )
        webhook_filter_check.grid(row=8, column=0, columnspan=2, sticky='w', padx=3, pady=3)

        ttk.Label(printer_frame, text="Min Value:").grid(row=9, column=0, sticky='w', padx=3, pady=3)
        webhook_filter_min_var = tk.StringVar(value=self.settings_config.get('printer_settings', {}).get('webhook_filter_min', '0'))
        webhook_filter_min_entry = ttk.Entry(printer_frame, textvariable=webhook_filter_min_var, width=10)
        webhook_filter_min_entry.grid(row=9, column=1, sticky='w', padx=3, pady=3)

        ttk.Label(printer_frame, text="Max Value:").grid(row=10, column=0, sticky='w', padx=3, pady=3)
        webhook_filter_max_var = tk.StringVar(value=self.settings_config.get('printer_settings', {}).get('webhook_filter_max', '100'))
        webhook_filter_max_entry = ttk.Entry(printer_frame, textvariable=webhook_filter_max_var, width=10)
        webhook_filter_max_entry.grid(row=10, column=1, sticky='w', padx=3, pady=3)

        save_printer_btn = ttk.Button(
            printer_frame,
            text="Save Printer Settings",
            command=lambda: self.save_printer_settings(
                printer_enabled_var, printer_baud_var, init_char_var, expected_response_var,
                command_var, end_char_var, reset_connection_var, webhook_filter_enabled_var,
                webhook_filter_min_var, webhook_filter_max_var
            )
        )
        save_printer_btn.grid(row=11, column=1, sticky='e', padx=3, pady=3)

        self.toggle_printer_settings(printer_enabled_var, printer_frame)

        # Close Button
        close_frame = ttk.Frame(frame)
        close_frame.pack(fill='x', pady=5)
        close_btn = ttk.Button(close_frame, text="Close", command=settings_win.destroy)
        close_btn.pack(side='right')

        self.update_settings_ui(offline_mode_var, webhook_frame, save_webhook_btn)

    def toggle_printer_settings(self, enabled_var, printer_frame):
        state = 'normal' if enabled_var.get() else 'disabled'
        for widget in printer_frame.winfo_children():
            if isinstance(widget, (ttk.Combobox, ttk.Entry, ttk.Checkbutton)) and widget != printer_frame.winfo_children()[0]:
                widget.config(state=state)

    def save_printer_settings(self, enabled_var, baud_var, init_char_var, expected_response_var, command_var, end_char_var, reset_connection_var, filter_enabled_var, filter_min_var, filter_max_var):
        self.settings_config['printer_settings'] = {
            'enabled': enabled_var.get(),
            'baudrate': baud_var.get(),
            'init_char': init_char_var.get(),
            'expected_response': expected_response_var.get(),
            'command': command_var.get(),
            'end_char': end_char_var.get(),
            'reset_connection': reset_connection_var.get(),
            'webhook_filter_enabled': filter_enabled_var.get(),
            'webhook_filter_min': filter_min_var.get(),
            'webhook_filter_max': filter_max_var.get()
        }
        self.save_config(self.settings_config_path, self.settings_config)
        messagebox.showinfo("Success", "Printer settings saved successfully.")

    ### HELPER: GET PORTS / SERIAL CONFIG
    def get_ports(self):
        ps = serial.tools.list_ports.comports()
        return [p.device for p in sorted(ps, key=lambda x: x.device)]

    def get_parity(self, parity_str):
        ps = parity_str.lower()
        return {
            'none': serial.PARITY_NONE,
            'even': serial.PARITY_EVEN,
            'odd': serial.PARITY_ODD,
            'mark': serial.PARITY_MARK,
            'space': serial.PARITY_SPACE
        }.get(ps, serial.PARITY_NONE)

    def get_stop_bits(self, stop_str):
        return {
            '1': serial.STOPBITS_ONE,
            '1.5': serial.STOPBITS_ONE_POINT_FIVE,
            '2': serial.STOPBITS_TWO
        }.get(stop_str, serial.STOPBITS_ONE)

    def get_flow_control(self, flow_str):
        fs = flow_str.lower()
        rtscts = (fs == 'rts/cts')
        xonxoff = (fs == 'xon/xoff')
        return (rtscts, xonxoff)

    ### SETTINGS / STARTUP
    def save_webhook_config(self, data_webhook_var, heartbeat_url_var, token_url_var):
        self.webhook_config['webhook_url'] = data_webhook_var.get()
        self.webhook_config['heartbeat_webhook_url'] = heartbeat_url_var.get()
        self.webhook_config['token_url'] = token_url_var.get()
        self.save_config(self.webhook_config_path, self.webhook_config)
        messagebox.showinfo("Success", "Webhook settings saved successfully.")

    def save_computer_name(self, comp_name_var):
        self.settings_config['computer_name'] = comp_name_var.get()
        self.save_config(self.settings_config_path, self.settings_config)
        messagebox.showinfo("Success", "Computer name saved successfully.")

    def toggle_launch_on_startup(self, var):
        self.settings_config['launch_on_startup'] = var.get()
        self.save_config(self.settings_config_path, self.settings_config)
        self.configure_launch_on_startup(var.get())

    def toggle_launch_minimized(self, var):
        self.settings_config['launch_minimized'] = var.get()
        self.save_config(self.settings_config_path, self.settings_config)

    def toggle_offline_mode(self, var, settings_win):
        offline_mode = var.get()
        self.settings_config['offline_mode'] = offline_mode
        self.save_config(self.settings_config_path, self.settings_config)
        self.heartbeat_flag = not offline_mode
        self.root.tk.call('wm', 'attributes', '.', '-topmost', '1')
        self.root.tk.call('wm', 'attributes', '.', '-topmost', '0')
        messagebox.showinfo(
            "Restart Required",
            "Please restart the application for offline mode changes to take full effect."
        )
        for tab in self.device_tabs:
            self.enforce_offline_mode(tab)
        self.update_settings_ui(
            var,
            settings_win.children['!notebook'].children['!frame'].children['!labelframe'],
            settings_win.children['!notebook'].children['!frame'].children['!labelframe'].children['!button3']
        )
        self.root.tk.call(
            self.root._w,
            'menu',
            'entryconfigure',
            'Connection',
            '-state',
            'disabled' if offline_mode else 'normal'
        )

    def update_settings_ui(self, offline_mode_var, webhook_frame, save_webhook_btn):
        state = 'disabled' if offline_mode_var.get() else 'normal'
        for widget in webhook_frame.winfo_children():
            if widget != save_webhook_btn:
                widget.configure(state=state)
        save_webhook_btn.configure(state=state)

    def enforce_offline_mode(self, tab):
        if self.settings_config.get('offline_mode', False):
            w = tab.widgets
            if w['data_type_var'].get() not in ['txt', 'printer']:
                w['data_type_var'].set('txt')
            w['handling_var'].set('cursor')
            for dt in w['data_type_radiobuttons']:
                w['data_type_radiobuttons'][dt].config(
                    state='disabled' if dt not in ['txt', 'printer'] else 'normal'
                )
            w['handling_radiobuttons']['webhook'].config(state='disabled')
            w['handling_radiobuttons']['cursor'].config(state='normal')
            w['token_entry'].config(state='disabled')
            w['token_lock_button'].config(state='disabled')
            self.toggle_folder_or_port_fields(tab)

    def configure_launch_on_startup(self, enable):
        os_name = platform.system()
        if os_name == 'Windows':
            self.configure_startup_windows(enable)
        elif os_name == 'Darwin':
            self.configure_startup_mac(enable)
        elif os_name == 'Linux':
            self.configure_startup_linux(enable)
        else:
            messagebox.showwarning("Unsupported OS", "Not supported on this OS.")

    def configure_startup_windows(self, enable):
        import winreg
        app_name = "SerialApp"
        exe_path = sys.executable if getattr(sys, 'frozen', False) else sys.argv[0]
        try:
            reg = winreg.ConnectRegistry(None, winreg.HKEY_CURRENT_USER)
            key_path = r"Software\Microsoft\Windows\CurrentVersion\Run"
            key = winreg.OpenKey(reg, key_path, 0, winreg.KEY_ALL_ACCESS)
            if enable:
                winreg.SetValueEx(key, app_name, 0, winreg.REG_SZ, exe_path)
            else:
                try:
                    winreg.DeleteValue(key, app_name)
                except FileNotFoundError:
                    pass
            winreg.CloseKey(key)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update startup: {e}")

    def configure_startup_mac(self, enable):
        plist_path = os.path.expanduser(f"~/Library/LaunchAgents/com.{getpass.getuser()}.serialapp.plist")
        exe_path = sys.executable if getattr(sys, 'frozen', False) else sys.argv[0]
        if enable:
            content = f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
<plist version="1.0">
   <dict>
       <key>Label</key>
       <string>com.{getpass.getuser()}.serialapp</string>
       <key>ProgramArguments</key>
       <array>
           <string>{exe_path}</string>
       </array>
       <key>RunAtLoad</key>
       <true/>
   </dict>
</plist>"""
            try:
                with open(plist_path, 'w') as f:
                    f.write(content)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create plist: {e}")
        else:
            try:
                if os.path.exists(plist_path):
                    os.remove(plist_path)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to remove plist: {e}")

    def configure_startup_linux(self, enable):
        autostart_dir = os.path.expanduser("~/.config/autostart")
        os.makedirs(autostart_dir, exist_ok=True)
        desktop_path = os.path.join(autostart_dir, "serialapp.desktop")
        exe_path = sys.executable if getattr(sys, 'frozen', False) else sys.argv[0]
        entry = f"""[Desktop Entry]
Type=Application
Exec={exe_path}
Hidden=false
NoDisplay=false
X-GNOME-Autostart-enabled=true
Name=SerialApp
Comment=Start SerialApp on login
"""
        if enable:
            try:
                with open(desktop_path, 'w') as f:
                    f.write(entry)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create autostart: {e}")
        else:
            try:
                if os.path.exists(desktop_path):
                    os.remove(desktop_path)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to remove autostart: {e}")

    ### DEVICE MGMT
    def add_device(self):
        dev_count = len(self.devices_config['devices']) + 1
        new_name = f"Device{dev_count}"
        new_file = f"device{dev_count}_config.json"
        self.devices_config['devices'].append({"name": new_name, "config_file": new_file})
        self.save_config(self.devices_config_path, self.devices_config)

        dev_tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(dev_tab, text=new_name)
        self.device_tabs.append(dev_tab)
        self.data_queues.append(queue.Queue())
        self.create_device_tab(dev_tab, dev_count - 1, {"name": new_name, "config_file": new_file})
        self.populate_remove_device_combo()

    def populate_remove_device_combo(self):
        names = [d["name"] for d in self.devices_config["devices"]]
        if hasattr(self, 'remove_combo'):
            self.remove_combo['values'] = names
            if names:
                self.remove_combo_var.set(names[0])

    def remove_selected_device(self):
        if len(self.devices_config["devices"]) <= 1:
            messagebox.showwarning("Warning", "At least 1 device must remain.")
            return

        selected_name = self.remove_combo_var.get()
        if not selected_name:
            return

        dev_index = next(
            (i for i, d in enumerate(self.devices_config['devices']) if d["name"] == selected_name),
            None
        )
        if dev_index is None:
            messagebox.showinfo("Info", "Could not find the selected device.")
            return

        confirm = messagebox.askyesno("Confirm Removal", f"Are you sure you want to remove '{selected_name}'?")
        if not confirm:
            return

        self.stop_device(dev_index)
        self.devices_config['devices'].pop(dev_index)
        self.save_config(self.devices_config_path, self.devices_config)

        tab_to_remove = self.device_tabs.pop(dev_index)
        self.notebook.forget(tab_to_remove)
        self.serial_ports.pop(dev_index)
        self.reading_threads.pop(dev_index)
        self.reading_flags.pop(dev_index)
        self.monitoring_threads.pop(dev_index)
        self.monitoring_flags.pop(dev_index)
        self.data_queues.pop(dev_index)
        self.update_status_bar()

    ### START/STOP
    def start_all_devices(self):
        for i in range(len(self.device_tabs)):
            self.start_device(i, on_startup=True)
        if not self.settings_config.get('offline_mode', False):
            self.start_heartbeat_thread()
        self.update_status_bar()

    def start_device(self, dev_index, on_startup=False):
        tab = self.device_tabs[dev_index]
        w = tab.widgets
        dt = w['data_type_var'].get()
        handling = w['handling_var'].get()
        folder = w['folder_path_var'].get()

        self.save_single_device_config(dev_index, quiet=True)

        dev_file = config_path(self.devices_config['devices'][dev_index]['config_file'])
        with open(dev_file, 'r') as f:
            dev_conf = json.load(f)

        if dt in ['txt', 'printer'] and handling == 'cursor':
            port = w['port_var'].get()
            if not port:
                if not on_startup:
                    messagebox.showerror("Error", "Select a port.")
                return
            if port in self.serial_ports_in_use:
                err = f"Error opening port {port}: Port in use."
                if not on_startup:
                    messagebox.showerror("Port Error", err)
                self.append_log(w['text_box'], err + "\n")
                return
            try:
                # Use printer-specific baud rate if enabled and data type is 'printer'
                baudrate = int(w['baud_var'].get())
                if dt == 'printer' and self.settings_config.get('printer_settings', {}).get('enabled', False):
                    baudrate = int(self.settings_config['printer_settings']['baudrate'])

                ser = serial.Serial(
                    port=port,
                    baudrate=baudrate,
                    parity=self.get_parity(w['parity_var'].get()),
                    bytesize=int(w['data_bits_var'].get()),
                    stopbits=self.get_stop_bits(w['stop_bits_var'].get()),
                    rtscts=self.get_flow_control(w['flow_control_var'].get())[0],
                    xonxoff=self.get_flow_control(w['flow_control_var'].get())[1],
                    timeout=float(w['interval_var'].get())
                )
                if w['flow_control_var'].get() == 'RTS/CTS':
                    ser.setRTS(True)
                elif w['flow_control_var'].get() == 'XON/XOFF':
                    ser.write(b'\x11')
                self.serial_ports[dev_index] = ser
                self.serial_ports_in_use.add(port)
                self.reading_flags[dev_index] = True
                self.last_data_time = time.time()
                t = threading.Thread(target=self.read_serial, args=(dev_index,))
                self.reading_threads[dev_index] = t
                t.start()

                self.lock_device_fields(tab)
                w['start_button'].config(state='disabled')
                w['stop_button'].config(state='normal')
                w['status_label'].config(text="Status: Running", foreground='green')
                logging.info(f"Device {w['device_name'].get()} started on {port}")
                self.update_status_bar()
            except Exception as e:
                err = f"Error opening port {port}: {e}"
                if not on_startup:
                    messagebox.showerror("Serial Port Error", err)
                self.append_log(w['text_box'], err + "\n")
                logging.error(err)
                return
        else:
            if not folder:
                if not on_startup:
                    messagebox.showerror("Error", "Select a folder to monitor.")
                return
            try:
                self.monitoring_flags[dev_index] = True
                t = threading.Thread(target=self.monitor_folder, args=(dev_index,))
                self.monitoring_threads[dev_index] = t
                t.start()
                self.webhook_threads.append(self.start_webhook_processor(dev_index))

                self.lock_device_fields(tab)
                w['start_button'].config(state='disabled')
                w['stop_button'].config(state='normal')
                w['status_label'].config(text="Status: Monitoring", foreground='green')
                self.update_status_bar()
            except Exception as e:
                err = f"Error starting monitor: {e}"
                if not on_startup:
                    messagebox.showerror("Monitoring Error", err)
                self.append_log(w['text_box'], err + "\n")
                return

    def stop_device(self, dev_index):
        tab = self.device_tabs[dev_index]
        w = tab.widgets
        dt = w['data_type_var'].get()
        handling = w['handling_var'].get()

        if dt in ['csv', 'xls', 'pdf'] or (dt in ['txt', 'printer'] and handling == 'webhook'):
            self.monitoring_flags[dev_index] = False
        else:
            self.reading_flags[dev_index] = False
            ser = self.serial_ports[dev_index]
            if ser and ser.is_open:
                port = ser.port
                ser.close()
                self.serial_ports_in_use.discard(port)
                self.serial_ports[dev_index] = None

        self.unlock_device_fields(tab)
        w['start_button'].config(state='normal')
        w['stop_button'].config(state='disabled')
        w['status_label'].config(text="Status: Disconnected", foreground='red')
        self.update_status_bar()

    def append_log(self, text_box, message):
        text_box.config(state='normal')
        text_box.insert('end', message)
        text_box.config(state='disabled')
        text_box.see('end')
        try:
            with open(self.main_log_file, 'a', encoding='utf-8') as f:
                f.write(message)
        except Exception as e:
            logging.error(f"Error writing to main log file: {e}")

    def clear_log(self, text_box):
        text_box.config(state='normal')
        text_box.delete('1.0', 'end')
        text_box.config(state='disabled')

    ### SERIAL READING
    def read_serial(self, dev_index):
        tab = self.device_tabs[dev_index]
        w = tab.widgets
        ser = self.serial_ports[dev_index]
        handling = w['handling_var'].get()
        dev_name = w['device_name'].get()
        dev_token = w['token_var'].get()
        dt = w['data_type_var'].get()
        interval = float(w['interval_var'].get())
        bytes_received = 0

        if dt == 'txt':
            while self.reading_flags[dev_index]:
                try:
                    # Passively read incoming data without sending any commands
                    data = ser.readline().decode('ascii', errors='ignore').strip()
                    if data:
                        self.process_buffer(dev_index, data)
                        bytes_received += len(data)
                        w['bytes_received_var'].set(f"Bytes: {bytes_received}")
                except Exception as e:
                    self.append_log(w['text_box'], f"Serial Error: {e}\n")
                    logging.error(f"Device {dev_name} serial error: {e}")
                time.sleep(interval)
        elif dt == 'printer' and self.settings_config.get('printer_settings', {}).get('enabled', False):
            ps = self.settings_config['printer_settings']
            init_char = ps.get('init_char', '@')
            expected_response = ps.get('expected_response', '$(O)$')
            command = ps.get('command', 'R1')
            end_char = ps.get('end_char', '#')
            reset_connection = ps.get('reset_connection', True)

            while self.reading_flags[dev_index]:
                try:
                    # Send initiation character for printer
                    ser.write(init_char.encode('ascii'))
                    response = ser.read(len(expected_response))
                    if response.decode('ascii') == expected_response:
                        self.append_log(w['text_box'], "Communication initiated successfully\n")
                    else:
                        self.append_log(w['text_box'], f"Unexpected response: {response}\n")
                        time.sleep(1)
                        continue

                    # Send command for printer
                    ser.write(command.encode('ascii'))

                    # Read until end character
                    response = ser.read_until(end_char.encode('ascii')).decode('ascii', errors='ignore')
                    if response.endswith(end_char):
                        data = response[:-len(end_char)]
                        self.process_buffer(dev_index, data)
                        bytes_received += len(data)
                        w['bytes_received_var'].set(f"Bytes: {bytes_received}")
                    else:
                        self.append_log(w['text_box'], f"Invalid response: {response}\n")

                    if reset_connection:
                        ser.close()
                        ser.open()
                    time.sleep(interval)
                except Exception as e:
                    self.append_log(w['text_box'], f"Serial Error: {e}\n")
                    logging.error(f"Device {dev_name} serial error: {e}")
                    time.sleep(1)

    def process_buffer(self, dev_index, data):
        tab = self.device_tabs[dev_index]
        w = tab.widgets
        handling = w['handling_var'].get()
        dev_name = w['device_name'].get()
        dev_token = w['token_var'].get()
        dt = w['data_type_var'].get()

        if not data.strip():
            return

        ts = datetime.datetime.now().isoformat()
        msg = f"{ts} - {data}\n"

        payload = {
            "name": dev_name,
            "token": dev_token,
            "time_sent": ts,
            "file_type": dt,
            "data": {"Sheet1": [{"A1": data}]}
        }

        if handling == 'webhook' and not self.settings_config.get('offline_mode', False):
            self.data_queues[dev_index].put(payload)
        elif handling == 'cursor':
            pyautogui.typewrite(data)

        self.append_log(w['text_box'], msg)

    def start_webhook_processor(self, dev_index):
        def process_queue():
            while self.reading_flags[dev_index] or self.monitoring_flags[dev_index]:
                try:
                    payload = self.data_queues[dev_index].get(timeout=1)
                    response = self.send_data_to_webhook(payload)
                    if response and response.status_code != 200:
                        self.append_log(
                            self.device_tabs[dev_index].widgets['text_box'],
                            f"Webhook failed: {response.status_code} - {response.text}\n"
                        )
                    self.data_queues[dev_index].task_done()
                except queue.Empty:
                    continue
                except Exception as e:
                    logging.error(f"Webhook processing error: {e}")
        thread = threading.Thread(target=process_queue, daemon=True)
        thread.start()
        return thread

    ### HEARTBEAT
    def start_heartbeat_thread(self):
        if not self.settings_config.get('offline_mode', False):
            t = threading.Thread(target=self.heartbeat_loop, daemon=True)
            t.start()

    def heartbeat_loop(self):
        while self.heartbeat_flag and not self.settings_config.get('offline_mode', False):
            for i in range(len(self.device_tabs)):
                self.send_device_heartbeat(i)
            time.sleep(30)

    def send_device_heartbeat(self, dev_index):
        if self.settings_config.get('offline_mode', False):
            return
        tab = self.device_tabs[dev_index]
        w = tab.widgets
        dev_name = w['device_name'].get()
        dt = w['data_type_var'].get()
        handling = w['handling_var'].get()
        dev_token = w['token_var'].get()

        local_ok = False
        if dt in ['txt', 'printer'] and handling == 'cursor':
            ser = self.serial_ports[dev_index]
            if ser and ser.is_open:
                local_ok = True
        else:
            folder = w['folder_path_var'].get()
            if folder and os.path.isdir(folder):
                try:
                    ext_map = {'txt': 'txt', 'csv': 'csv', 'xls': 'xls', 'pdf': 'pdf', 'printer': 'prn'}
                    ext = ext_map.get(dt, 'txt')
                    hb_file = os.path.join(folder, f"heartbeat_test.{ext}")
                    with open(hb_file, 'w') as f:
                        f.write("heartbeat test")
                    os.remove(hb_file)
                    local_ok = True
                except:
                    pass

        with self.token_lock:
            if not self.token or time.time() > self.token_expiry:
                self.request_token()
        if not self.token:
            w['heartbeat_indicator'].config(bg='red')
            return
        hb_url = self.webhook_config.get('heartbeat_webhook_url', '')
        headers = {
            **self.webhook_config.get('webhook_headers', {}),
            "Authorization": f"Bearer {self.token}"
        }
        payload = {
            "computer_name": self.settings_config.get('computer_name', "MyComputer"),
            "device_name": dev_name,
            "token": dev_token,
            "time_tested": datetime.datetime.now().isoformat(),
            "connected": local_ok
        }
        try:
            r = requests.post(hb_url, headers=headers, json=payload, timeout=5)
            if r.status_code == 200:
                w['heartbeat_indicator'].config(bg='green')
            else:
                w['heartbeat_indicator'].config(bg='red')
        except Exception as e:
            logging.error(f"Error sending heartbeat: {e}")
            w['heartbeat_indicator'].config(bg='red')

    ### FOLDER MONITOR
    def monitor_folder(self, dev_index):
        tab = self.device_tabs[dev_index]
        w = tab.widgets
        folder = w['folder_path_var'].get()

        complete_dir = os.path.join(folder, 'Complete')
        pending_dir = os.path.join(folder, 'Pending')
        failed_dir = os.path.join(folder, 'Failed')
        for d in [complete_dir, pending_dir, failed_dir]:
            os.makedirs(d, exist_ok=True)

        retry_counts = {}
        while self.monitoring_flags[dev_index]:
            self.process_files_in_folder(
                tab,
                folder,
                retry_counts,
                pending_dir,
                failed_dir,
                complete_dir,
                is_pending=False
            )
            self.process_files_in_folder(
                tab,
                pending_dir,
                retry_counts,
                pending_dir,
                failed_dir,
                complete_dir,
                is_pending=True
            )
            time.sleep(1)

    def move_file_with_retry(self, src, dest, max_retries=5, delay=2):
        for attempt in range(max_retries):
            try:
                os.rename(src, dest)
                return True, None
            except Exception as e:
                logging.warning(f"Move attempt {attempt + 1}/{max_retries} failed for {src}: {e}")
                if attempt < max_retries - 1:
                    time.sleep(delay)
                else:
                    return False, str(e)
        return False, "Unexpected error"

    def process_files_in_folder(self, tab, folder_path, retry_counts, pending_dir, failed_dir, complete_dir, is_pending=False):
        w = tab.widgets
        dt = w['data_type_var'].get()
        handling_method = w['handling_var'].get()
        dev_name = w['device_name'].get()
        dev_token = w['token_var'].get()

        if not os.path.isdir(folder_path):
            return

        files = os.listdir(folder_path)
        for file in files:
            if not self.monitoring_flags[self.device_tabs.index(tab)]:
                break

            file_path = os.path.join(folder_path, file)
            if os.path.isdir(file_path) or file.lower().startswith("heartbeat_test."):
                continue

            ext = os.path.splitext(file.lower())[-1].lstrip('.')
            valid_exts = {
                'csv': ['csv'],
                'pdf': ['pdf'],
                'xls': ['xls', 'xlsx', 'xml'],
                'txt': ['txt'],
                'printer': ['prn', 'txt']
            }.get(dt, [])
            if ext not in valid_exts:
                continue

            # Skip files that were modified recently
            if os.path.getmtime(file_path) > time.time() - 5:
                continue

            now = time.time()
            last_attempt = retry_counts.get(file, {}).get('last_attempt', 0)
            attempts = retry_counts.get(file, {}).get('count', 0)

            if is_pending and (now - last_attempt < 30):
                continue

            success = False
            reason = ""
            try:
                parsed_sheets = self.parse_file(file_path, ext)
            except Exception as e:
                reason = str(e)
                parsed_sheets = {}

            # Add a small delay to ensure file handle is released
            time.sleep(1)

            ts = datetime.datetime.now().isoformat()
            payload = {
                "name": dev_name,
                "token": dev_token,
                "time_sent": ts,
                "file_name": file,
                "file_type": ext,
                "data": parsed_sheets
            }

            if handling_method == 'webhook':
                send_to_webhook = True
                if dt == 'printer' and self.settings_config.get('printer_settings', {}).get('enabled', False):
                    ps = self.settings_config['printer_settings']
                    if ps.get('webhook_filter_enabled', False):
                        try:
                            value_str = parsed_sheets["Sheet1"][0]["A1"]
                            value = float(value_str)
                            min_val = float(ps['webhook_filter_min'])
                            max_val = float(ps['webhook_filter_max'])
                            if min_val <= value <= max_val:
                                send_to_webhook = False
                                reason = f"Value {value} within filter range [{min_val}, {max_val}], not sending."
                                self.append_log(w['text_box'], reason + "\n")
                        except (ValueError, KeyError):
                            pass

                if send_to_webhook:
                    resp = self.send_data_to_webhook(payload)
                    success = resp and resp.status_code == 200
                    reason = "No response" if resp is None else f"{resp.status_code} {resp.text}" if not success else ""
                else:
                    success = True  # Filtered out
            elif handling_method == 'cursor':
                data_str = json.dumps(payload, ensure_ascii=False)
                pyautogui.typewrite(data_str)
                success = True

            if success:
                dest = os.path.join(complete_dir, file)
                if os.path.exists(dest):
                    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                    name, ext_with_dot = os.path.splitext(file)
                    new_file = f"{name}_{timestamp}{ext_with_dot}"
                    dest = os.path.join(complete_dir, new_file)
                    self.append_log(w['text_box'], f"File name conflict: Renamed '{file}' to '{new_file}'\n")

                move_success, move_error = self.move_file_with_retry(file_path, dest)
                if move_success:
                    if file in retry_counts:
                        del retry_counts[file]
                    self.append_log(w['text_box'], f"Processed file: {file}\n")
                else:
                    attempts += 1
                    retry_counts[file] = {'count': attempts, 'last_attempt': now}
                    self.append_log(w['text_box'], f"Failed to move {file} to Complete: {move_error}\n")
                    dest = os.path.join(pending_dir, file)
                    move_success, move_error = self.move_file_with_retry(file_path, dest)
                    if move_success:
                        self.append_log(
                            w['text_box'],
                            f"Moved {file} to Pending for retry (attempt {attempts}/5)\n"
                        )
                    else:
                        self.append_log(
                            w['text_box'],
                            f"Could not move {file} to Pending: {move_error}. Keeping in original location.\n"
                        )
            else:
                attempts += 1
                retry_counts[file] = {'count': attempts, 'last_attempt': now}
                if attempts >= 5:
                    dest = os.path.join(failed_dir, file)
                    move_success, move_error = self.move_file_with_retry(file_path, dest)
                    if move_success:
                        self.append_log(
                            w['text_box'],
                            f"File {file} permanently failed after 5 attempts. Reason: {reason}\nMoved to Failed.\n"
                        )
                    else:
                        self.append_log(
                            w['text_box'],
                            f"File {file} failed after 5 attempts. Reason: {reason}\nCould not move to Failed: {move_error}\n"
                        )
                else:
                    dest = os.path.join(pending_dir, file)
                    move_success, move_error = self.move_file_with_retry(file_path, dest)
                    if move_success:
                        self.append_log(
                            w['text_box'],
                            f"File {file} failed (attempt {attempts}/5). Reason: {reason}\nMoved to Pending.\n"
                        )
                    else:
                        self.append_log(
                            w['text_box'],
                            f"File {file} failed (attempt {attempts}/5). Reason: {reason}\nCould not move to Pending: {move_error}\n"
                        )

    ### TOKEN LOCK
    def toggle_token_lock(self, dev_index):
        tab = self.device_tabs[dev_index]
        w = tab.widgets
        locked = w['token_locked']
        w['token_locked'] = not locked
        w['token_entry'].config(state='disabled' if not locked else 'normal')
        w['token_lock_button'].config(text="🔒" if not locked else "🔓")

    ### HANDLING / FOLDER OR PORT
    def update_handling_options(self, tab):
        dt = tab.widgets['data_type_var'].get()
        handling_var = tab.widgets['handling_var']
        handling_rbs = tab.widgets['handling_radiobuttons']

        if dt in ['csv', 'xls', 'pdf']:
            handling_rbs['cursor'].config(state='disabled')
            if handling_var.get() == 'cursor':
                handling_var.set('webhook')
            handling_rbs['webhook'].config(state='normal')
        else:  # 'txt' or 'printer'
            if self.settings_config.get('offline_mode', False):
                handling_rbs['cursor'].config(state='normal')
                handling_rbs['webhook'].config(state='disabled')
                if handling_var.get() == 'webhook':
                    handling_var.set('cursor')
            else:
                handling_rbs['cursor'].config(state='normal')
                handling_rbs['webhook'].config(state='normal')
        self.toggle_folder_or_port_fields(tab)

    def update_data_type_options(self, tab):
        handling = tab.widgets['handling_var'].get()
        dt_var = tab.widgets['data_type_var']
        dt_rbs = tab.widgets['data_type_radiobuttons']

        if handling == 'cursor' or self.settings_config.get('offline_mode', False):
            if dt_var.get() not in ['txt', 'printer']:
                dt_var.set('txt')
            for dt_ in dt_rbs:
                dt_rbs[dt_].config(state='disabled' if dt_ not in ['txt', 'printer'] else 'normal')
        else:
            for dt_ in dt_rbs:
                dt_rbs[dt_].config(state='normal')
        self.toggle_folder_or_port_fields(tab)

    def toggle_folder_or_port_fields(self, tab):
        w = tab.widgets
        dt = w['data_type_var'].get()
        hm = w['handling_var'].get()

        if dt in ['txt', 'printer'] and (
            hm == 'cursor' or (
                self.settings_config.get('offline_mode', False) and hm != 'webhook'
            )
        ):
            for key in [
                'port_combo',
                'baud_combo',
                'parity_combo',
                'data_bits_combo',
                'stop_bits_combo',
                'flow_control_combo',
                'interval_entry'
            ]:
                w[key].config(state='normal')
            w['folder_entry'].config(state='disabled')
            w['browse_button'].config(state='disabled')
        else:
            for key in [
                'port_combo',
                'baud_combo',
                'parity_combo',
                'data_bits_combo',
                'stop_bits_combo',
                'flow_control_combo',
                'interval_entry'
            ]:
                w[key].config(state='disabled')
            w['folder_entry'].config(state='normal')
            w['browse_button'].config(state='normal')

    def browse_folder(self, folder_path_var):
        sel = filedialog.askdirectory()
        if sel:
            folder_path_var.set(sel)

    ### SEND TO WEBHOOK
    def request_token(self):
        token_url = self.webhook_config.get('token_url', '')
        if not token_url:
            logging.warning("Token URL not configured; cannot request token")
            self.token = None
            self.token_expiry = 0
            return

        license_data = self.load_license()
        license_key = license_data.get("license_key", "")
        logging.debug(f"Using license_key from license.json: {license_key}")  # Add this for debugging
        payload = {"license_key": license_key}
        mac_address = self.get_mac_address()
        if mac_address:
            payload["mac_address"] = mac_address
        serial_number = self.get_serial_number()
        if serial_number:
            payload["serial_number"] = serial_number

        try:
            r = requests.post(token_url, json=payload, timeout=5)
            if r.status_code == 200:
                data = r.json()
                self.token = data["token"]
                self.token_expiry = time.time() + data.get("expires_in", 3600)
                logging.info(f"Token retrieved: {self.token}")
            else:
                logging.error(f"Failed to retrieve token: {r.status_code} - {r.text}")
                self.token = None
                self.token_expiry = 0
        except Exception as e:
            logging.error(f"Error requesting token: {e}")
            self.token = None
            self.token_expiry = 0

    def send_data_to_webhook(self, payload):
        if self.settings_config.get('offline_mode', False):
            return None
        with self.token_lock:
            if not self.token or time.time() > self.token_expiry:
                self.request_token()
        if not self.token:
            logging.error("No valid token available")
            return None
        url = self.webhook_config.get('webhook_url', '')
        headers = {
            **self.webhook_config.get('webhook_headers', {}),
            "Authorization": f"Bearer {self.token}"
        }
        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=10)
            return resp
        except Exception as e:
            logging.error(f"Error sending data to webhook: {e}")
            return None

    ### SAVE ALL
    def save_all_device_configs(self):
        for i in range(len(self.device_tabs)):
            self.save_single_device_config(i, quiet=True)
        messagebox.showinfo("Success", "All device settings saved.")

    ### PARSE FILE
    def parse_file(self, file_path, ext):
        def xml_to_dict(element):
            if len(element) == 0:
                return element.text
            result = {}
            for child in element:
                child_dict = xml_to_dict(child)
                if child.tag in result:
                    if isinstance(result[child.tag], list):
                        result[child.tag].append(child_dict)
                    else:
                        result[child.tag] = [result[child.tag], child_dict]
                else:
                    result[child.tag] = child_dict
            return result

        sheets_dict = {}
        ext = file_path.rsplit('.', 1)[-1].lower()

        def cell_name(r, c):
            col_letters = ""
            temp_c = c
            while True:
                temp_c, remainder = divmod(temp_c, 26)
                col_letters = chr(65 + remainder) + col_letters
                if temp_c == 0:
                    break
                temp_c -= 1
            return f"{col_letters}{r+1}"

        try:
            if ext in ['txt', 'prn']:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    raw_data = f.read()
                clean_data = ''.join(ch for ch in raw_data if ch.isprintable())
                sheets_dict["Sheet1"] = [{"A1": clean_data}]
            elif ext == 'csv':
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    all_rows = list(reader)
                sheets_dict["Sheet1"] = [
                    {cell_name(r_idx, c_idx): val for c_idx, val in enumerate(row)}
                    for r_idx, row in enumerate(all_rows)
                ]
            elif ext in ['xls', 'xlsx', 'xml']:
                if ext == 'xls':
                    book = xlrd.open_workbook(file_path)
                    for sh in book.sheets():
                        sdata = []
                        for r in range(sh.nrows):
                            row_dict = {}
                            for c in range(sh.ncols):
                                cell_val = sh.cell_value(r, c)
                                ctype = sh.cell_type(r, c)
                                if ctype == xlrd.XL_CELL_DATE:
                                    cell_val = xlrd.xldate.xldate_as_datetime(
                                        cell_val, book.datemode
                                    ).isoformat()
                                else:
                                    cell_val = str(cell_val)
                                row_dict[cell_name(r, c)] = cell_val
                            sdata.append(row_dict)
                        sheets_dict[sh.name] = sdata
                    book.release_resources()  # Ensure resources are released
                elif ext == 'xlsx':
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    for sh_name in wb.sheetnames:
                        ws = wb[sh_name]
                        sdata = [
                            {
                                cell_name(r_idx, c_idx): str(cell_val)
                                if cell_val is not None
                                else ""
                                if not isinstance(cell_val, datetime.datetime)
                                else cell_val.isoformat()
                                for c_idx, cell_val in enumerate(row)
                            }
                            for r_idx, row in enumerate(ws.iter_rows(values_only=True))
                        ]
                        sheets_dict[sh_name] = sdata
                    wb.close()  # Explicitly close the workbook

                    # Extract images from XLSX (ZIP archive)
                    with zipfile.ZipFile(file_path, 'r') as zf:
                        image_list = [f for f in zf.namelist() if f.startswith('xl/media/')]
                        images = []
                        for img_path in image_list:
                            with zf.open(img_path) as img_file:
                                img_data = img_file.read()
                                img_format = img_path.rsplit('.', 1)[-1].lower()
                                base64_image = base64.b64encode(img_data).decode('utf-8')
                                images.append({"data": base64_image, "format": img_format})
                        if images:
                            # Add images to the first sheet (adjust as needed)
                            sheets_dict[list(sheets_dict.keys())[0]]["images"] = images
                else:  # xml
                    tree = ET.parse(file_path)
                    root = tree.getroot()
                    sheets_dict["Sheet1"] = [xml_to_dict(root)]
            elif ext == 'pdf':
                doc = fitz.open(file_path)
                for page_num in range(len(doc)):
                    page = doc[page_num]
                    text = page.get_text("text")  # Extract text for the page
                    images = []
                    for img in page.get_images(full=True):
                        xref = img[0]
                        base_image = doc.extract_image(xref)
                        image_bytes = base_image["image"]
                        image_format = base_image["ext"]
                        base64_image = base64.b64encode(image_bytes).decode('utf-8')
                        images.append({"data": base64_image, "format": image_format})
                    sheets_dict[f"Page {page_num + 1}"] = {"text": text, "images": images}
                doc.close()
            else:
                with open(file_path, 'rb') as f:
                    sheets_dict["Sheet1"] = [{"A1": str(f.read())}]
        except Exception as e:
            logging.error(f"Error parsing file {file_path}: {e}")
            raise e
        return sheets_dict

    ### LOCK/UNLOCK DEVICE FIELDS
    def lock_device_fields(self, tab):
        w = tab.widgets
        w['device_name_entry'].config(state='disabled')
        for rb in w['data_type_radiobuttons'].values():
            rb.config(state='disabled')
        for rb in w['handling_radiobuttons'].values():
            rb.config(state='disabled')
        w['folder_entry'].config(state='disabled')
        w['browse_button'].config(state='disabled')
        w['token_entry'].config(state='disabled')
        w['token_lock_button'].config(state='disabled')
        w['port_combo'].config(state='disabled')
        w['baud_combo'].config(state='disabled')
        w['parity_combo'].config(state='disabled')
        w['data_bits_combo'].config(state='disabled')
        w['stop_bits_combo'].config(state='disabled')
        w['flow_control_combo'].config(state='disabled')
        w['interval_entry'].config(state='disabled')

    def unlock_device_fields(self, tab):
        w = tab.widgets
        w['device_name_entry'].config(state='normal')
        for rb in w['data_type_radiobuttons'].values():
            rb.config(state='normal')
        for rb in w['handling_radiobuttons'].values():
            rb.config(state='normal')
        self.update_handling_options(tab)
        w['token_entry'].config(state='disabled' if w['token_locked'] else 'normal')
        w['token_lock_button'].config(state='normal', text="🔒" if w['token_locked'] else "🔓")
        if self.settings_config.get('offline_mode', False):
            self.enforce_offline_mode(tab)

    ### NEW FEATURES
    def export_logs(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if filename:
            try:
                with open(self.main_log_file, 'r') as source, open(filename, 'w') as target:
                    target.write(source.read())
                messagebox.showinfo("Success", "Logs exported successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export logs: {e}")

    def backup_config(self):
        folder = filedialog.askdirectory(title="Select Backup Directory")
        if folder:
            try:
                config_dir = os.path.expanduser("~/.serial_app")
                backup_name = f"serial_app_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
                backup_path = os.path.join(folder, backup_name)
                shutil.copytree(config_dir, backup_path)
                messagebox.showinfo("Success", f"Configuration backed up to {backup_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to backup config: {e}")

    def restore_config(self):
        folder = filedialog.askdirectory(title="Select Backup Directory")
        if folder:
            confirm = messagebox.askyesno("Confirm Restore", "This will overwrite current configs. Continue?")
            if confirm:
                try:
                    config_dir = os.path.expanduser("~/.serial_app")
                    shutil.rmtree(config_dir)
                    shutil.copytree(folder, config_dir)
                    messagebox.showinfo("Success", "Configuration restored. Restarting app...")
                    self.on_closing()
                    os.execl(sys.executable, sys.executable, *sys.argv)
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to restore config: {e}")

    def refresh_ports_periodically(self):
        for i, tab in enumerate(self.device_tabs):
            w = tab.widgets
            if not self.reading_flags[i]:
                w['port_combo']['values'] = self.get_ports()
        self.root.after(5000, self.refresh_ports_periodically)

    def update_status_bar(self):
        active = sum(1 for flag in self.reading_flags if flag) + sum(
            1 for flag in self.monitoring_flags if flag
        )
        self.status_var.set(f"Active Devices: {active}")

    ### CLEAN EXIT
    def on_closing(self):
        self.heartbeat_flag = False
        for i in range(len(self.device_tabs)):
            self.stop_device(i)
        for thread in self.reading_threads + self.monitoring_threads + self.webhook_threads:
            if thread and thread.is_alive():
                thread.join(timeout=1)
        for obs in self.observers:
            obs.stop()
            obs.join()
        self.root.destroy()
        logging.info("Application closed")

if __name__ == "__main__":
    root = ThemedTk(theme="adapta")
    app = SerialApp(root)
    root.mainloop()